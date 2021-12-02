import openpyxl as xl
from openpyxl.chart import BarChart, Reference

"""
wb = xl.load_workbook('data.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
cell = sheet.cell(1,1)

print('cell :', cell.value)

#print('sheet max_row :', sheet.max_row)

for row in range(1, sheet.max_row+1):
    #print('row: ', row)
    cell = sheet.cell(row, 3)
    print(cell.value)
"""
"""
wb = xl.load_workbook('data.xlsx')
sheet = wb['Sheet1'] # sheet
cell = sheet['a1'] # a = A column and 1 = one row
cell = sheet.cell(1,1) # one row | one column

print('cell ', cell.value)
print('max column: ', sheet.max_column)
print('max row: ', sheet.max_row)

# looping
for row in range(2, sheet.max_row + 1): # all collect row data
    #print(row)
    cell = sheet.cell(row, 3) # 4 rows and 3 column are collect
    print(cell.value)  # show row and 3 column
    print()

    total_row = cell.value * 0.9 # row data (price * 0.9 ) each value
    total_column = sheet.cell(row,4) # create new 4 column and entry data
    total_column.value = total_row # replace column


values = Reference(sheet,
                   min_row=2,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('data1.xlsx')
"""

def proces_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1'] # sheet


    # looping
    for row in range(2, sheet.max_row + 1): # all collect row data
        #print(row)
        cell = sheet.cell(row, 3) # 4 rows and 3 column are collect
        print(cell.value)  # show row and 3 column
        print()

        total_row = cell.value * 0.9 # row data (price * 0.9 ) each value
        total_column = sheet.cell(row,4) # create new 4 column and entry data
        total_column.value = total_row # replace column


    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)